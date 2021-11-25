from tkinter import *
from openpyxl import load_workbook
from datetime import datetime
from tkinter.messagebox import *
from os import startfile


class bill_menu:
    def __init__(self, file_data):
        self.Win_bill_menu = Tk()
        self.nxt_bill = Button(self.Win_bill_menu, text="Next Bill")
        self.close_win = Button(
            self.Win_bill_menu, text="close", command=lambda: self.Win_bill_menu.destroy())
        self.file_data = file_data
        self.menu = menu_bar(self.Win_bill_menu, self.file_data, bill_menu)
        self.billed_saree = []
        now = datetime.now()
        bill_no = now.strftime("%d%m%y")
        self.bill_no = bill_no
        self.today = now.strftime("%d-%m-%y")
        # ----------- Working in data from database (.xlsx file)-----------
        self.workbook = load_workbook(self.file_data)
        self.WHOLE_DATA = self.workbook["total data"]
        self.DAILY_DATA = self.workbook["daily sales"]
        self.SAREE_NO_FROM_DATABASE = [
            str(a.value) for a in self.WHOLE_DATA["A"]]
        self.UNSOLD_SAREE_NO_FROM_DATABASE = [str(self.WHOLE_DATA["A" + str(a + 1)].value) for a in
                                              range(1, len(self.SAREE_NO_FROM_DATABASE) + 1) if
                                              self.WHOLE_DATA["H" + str(a + 1)].value is None]
        print(self.UNSOLD_SAREE_NO_FROM_DATABASE)
        # --------done------
        bill_nos = [a.value for a in self.DAILY_DATA["A"]][1:]
        bill_nos = [int(a[6:]) for a in bill_nos if a[:6] == self.bill_no]
        # bill_nos = [a.value[6:] for a in self.DAILY_DATA["A"] if a.value[:6] == self.bill_no]
        if len(bill_nos) == 0:
            self.bill_no = self.bill_no + "01"
        else:
            if max(bill_nos) + 1 < 10:
                self.bill_no += "0" + str(max(bill_nos) + 1)
            else:
                self.bill_no += str(max(bill_nos) + 1)
        self.total = 0
        Label(self.Win_bill_menu, text="ABIRAMI SILKS", font=(
            "ARIAL", 45)).grid(column=0, row=0, columnspan=10)
        Label(self.Win_bill_menu, text="Name", font=(
            "ARIAL", 18)).grid(column=0, row=1)
        Label(self.Win_bill_menu, text="Number",
              font=("ARIAL", 18)).grid(column=2, row=1)
        Label(self.Win_bill_menu, text="Address",
              font=("ARIAL", 18)).grid(column=0, row=2)
        self.saree_num_button = Button(self.Win_bill_menu, text="Saree no", font=("ARIAL", 15),
                                       command=lambda: self.add_saree_to_bill(True))
        self.saree_num_button.grid(column=0, row=3)

        self.name_entry = Entry(
            self.Win_bill_menu, justify=CENTER, bd=2, font=("ARIAL", 18))
        self.name_entry.grid(column=1, row=1)

        self.number_entry = Entry(
            self.Win_bill_menu, justify=CENTER, bd=2, font=("ARIAL", 18))
        self.number_entry.grid(column=3, row=1)

        self.address_entry = Entry(
            self.Win_bill_menu, justify=CENTER, bd=2, width=50, font=("ARIAL", 18))
        self.address_entry.grid(column=1, row=2, columnspan=5)

        self.saree_num_entry = Entry(
            self.Win_bill_menu, justify=CENTER, bd=3, font=("ARIAL", 15))
        self.saree_num_entry.grid(column=1, row=3)
        self.saree_num_entry.bind("<Return>", self.add_saree_to_bill)

        Label(self.Win_bill_menu, text="SNO", width=10, font=(
            "ARIAL", 11, "bold")).grid(row=4, column=0)
        Label(self.Win_bill_menu, text="SAREE NO", width=10,
              font=("ARIAL", 11, "bold")).grid(row=4, column=1)
        Label(self.Win_bill_menu, text="DESIGN", width=10,
              font=("ARIAL", 11, "bold")).grid(row=4, column=2)
        Label(self.Win_bill_menu, text="COLOUR", width=10,
              font=("ARIAL", 11, "bold")).grid(row=4, column=3)
        Label(self.Win_bill_menu, text="PRICE", width=10,
              font=("ARIAL", 11, "bold")).grid(row=4, column=4)

        self.total_label = Label(
            self.Win_bill_menu, text=f"TOTAL={str(self.total)}", font=("ARIAL", 11))

        self.bill_button = Button(text="BILL", font=(
            "comforta", 11, "bold"), command=self.bill)
        self.Win_bill_menu.mainloop()

    def add_saree_to_bill(self, event):
        saree_nos = self.saree_num_entry.get()
        saree_nos = saree_nos.split("-")
        # to clear the saree_num_entry widget
        self.saree_num_entry.delete(0, END)
        for saree_no in saree_nos:
            if saree_no in self.UNSOLD_SAREE_NO_FROM_DATABASE and saree_no != "":
                row_no = self.SAREE_NO_FROM_DATABASE.index(saree_no) + 1
                sno = len(self.billed_saree) + 1
                colour = self.WHOLE_DATA["E" + str(row_no)].value
                design = self.WHOLE_DATA["D" + str(row_no)].value
                price = Entry(self.Win_bill_menu, font=(
                    "ARIAL", 11), justify=CENTER)
                price.grid(column=4, row=int(sno) + 4)
                saree_detail = [sno, saree_no, design,
                                colour, str(row_no), price]
                # --------to check colour is not none
                if self.WHOLE_DATA["E" + str(row_no)].value is None:
                    colour = Entry(self.Win_bill_menu)
                    colour.grid(column=3, row=4 + int(sno))
                    saree_detail[3] = colour
                else:
                    Label(self.Win_bill_menu, text=colour, width=10, font=("ARIAL", 11)).grid(column=3,
                                                                                              row=4 + int(sno))
                Label(self.Win_bill_menu, text=sno, width=10, font=(
                    "ARIAL", 11)).grid(column=0, row=4 + int(sno))
                Label(self.Win_bill_menu, text=saree_no, width=10, font=(
                    "ARIAL", 11)).grid(column=1, row=4 + int(sno))
                Label(self.Win_bill_menu, text=design, width=10, font=(
                    "ARIAL", 11)).grid(column=2, row=4 + int(sno))
                self.bill_button.grid(column=5, row=int(sno) + 5)
                self.total_label.config(text=f"TOTAL={str(self.total)}")
                self.total_label.grid(column=2, row=int(sno) + 5)
                self.billed_saree += [saree_detail]
                self.UNSOLD_SAREE_NO_FROM_DATABASE.remove(saree_no)
            else:
                # billed_saree structure [sno, saree_no, design, colour,str(row_no)]
                if saree_no != "":
                    showerror(title="wrong saree no",
                              message=f"you entered wrong saree no or sold saree no :{saree_no}")

    def nxt_bill_func(self):
        self.Win_bill_menu.destroy()
        bill_menu(self.file_data)

    def bill(self):
        # ----getting prices for all saree and getting colour for unmentioned saree
        # billed_saree structure [sno, saree_no, design, colour,str(row_no), price]
        for saree in self.billed_saree:
            price = saree[-1]
            if type(price) != (int or float):
                try:
                    if price.get() != "":
                        try:
                            price = float(price.get())
                            saree[-1].destroy()
                            Label(self.Win_bill_menu, text=price, width=10, font=("ARIAL", 11)).grid(column=4,
                                                                                                     row=4 + int(
                                                                                                         saree[0]))
                            self.total += price
                        except ValueError:
                            showwarning(title="Enter the price correctly",
                                        message=f"you  did not give the price of saree is not an number")
                            saree[-1].delete(0, END)
                except:
                    price = saree[-1]
            elif type(price) == (int or float):
                price = saree[-1]
            else:
                # billed_saree structure [sno, saree_no, design, colour,str(row_no), price]
                showerror(title="Enter the price",
                          message=f"you  did not give the price of saree(saree no:{saree[1]})")
            saree[-1] = price
            # ---getting colour for the unmentioned saree colour
            colour = saree[3]
            try:
                if type(colour) != str:
                    if colour.get() != "":
                        colour = colour.get()
                        saree[3].destroy()
                        Label(self.Win_bill_menu, text=colour, width=10, font=("ARIAL", 11)).grid(column=3,
                                                                                                  row=4 + int(saree[0]))
                        self.WHOLE_DATA["H" + saree[-2]] = colour
                        saree[3] = colour
                        print(saree)
                elif type(colour) == str:
                    saree[3] = colour
                else:
                    # billed_saree structure [sno, saree_no, design, colour,str(row_no), price]
                    showerror(title="Enter the colour",
                              message=f"you  did not give the colour of saree(saree no:{saree[1]})")
            except AttributeError:
                pass
        # billed_saree structure [sno, saree_no, design, colour,str(row_no)]

        for one_saree in self.billed_saree:
            # billed_saree structure [sno, saree_no, design, colour,str(row_no)]
            # -----------add data to the xlsx sheet cus name,num,address
            # A:'SAREE NO',B:'Date',C:'Name',D:'Design',E:'Colours',F:'Purchase Amount'	G:'DATE OF SALE' H:'SOLD PRICE'
            # I:'CUSTOMER NAME'	J:'CUSTOMER NUMBER'	k:'ADDRESS'	L:'BILL NO'

            self.WHOLE_DATA["E" + one_saree[-2]
                            ] = one_saree[3]  # customer name
            self.WHOLE_DATA["G" + one_saree[-2]] = self.today  # d.o.s
            self.WHOLE_DATA["H" + one_saree[-2]] = one_saree[-1]  # sold price
            self.WHOLE_DATA["I" + one_saree[-2]
                            ] = self.name_entry.get()  # customer name
            self.WHOLE_DATA["J" + one_saree[-2]
                            ] = self.number_entry.get()  # customer number
            self.WHOLE_DATA["K" + one_saree[-2]
                            ] = self.address_entry.get()  # customer address
            self.WHOLE_DATA["L" + one_saree[-2]] = self.bill_no  # bill no
            # -------------done in first sheet
            # --------second sheet "daily"
            # ---------- billed_saree structure [sno, saree_no, design, colour,str(row_no),price]
            # A:'BILL NO',B:'SAREE NO',C:'CUSTOMER NAME',D:'DATE OF SALE',E:'DESIGN',F:'COLOUR',G:'SOLD PRICE'
            row_ = str(len(self.DAILY_DATA["A"]) + 1)
            self.DAILY_DATA["A" + row_] = self.bill_no
            self.DAILY_DATA["B" + row_] = one_saree[1]
            self.DAILY_DATA["C" + row_] = self.name_entry.get()
            self.DAILY_DATA["D" + row_] = self.today
            self.DAILY_DATA["E" + row_] = one_saree[2]
            self.DAILY_DATA["F" + row_] = one_saree[3]
            self.DAILY_DATA["G" + row_] = one_saree[-1]
            Label(self.Win_bill_menu, text=self.name_entry.get(), width=10, font=("ARIAL", 11, "bold")).grid(
                column=1, row=1)
            Label(self.Win_bill_menu, text=self.number_entry.get(), width=10, font=("ARIAL", 11, "bold")).grid(
                column=3,
                row=1)
            Label(self.Win_bill_menu, text=self.address_entry.get(), width=10, font=("ARIAL", 11, "bold")).grid(
                column=1,
                row=2)

        self.name_entry.destroy()
        self.address_entry.destroy()
        self.number_entry.destroy()
        self.bill_button.destroy()
        self.saree_num_entry.destroy()
        self.saree_num_button.destroy()
        self.nxt_bill["command"] = self.nxt_bill_func
        self.DAILY_DATA["H" + str(len(self.DAILY_DATA["A"]))] = self.total
        print(self.total, self.bill_no)
        self.total_label.config(text=f"TOTAL={str(self.total)}")
        self.workbook.save(self.file_data)
        kk = self.billed_saree[-1][0]
        self.nxt_bill.grid(column=6, row=kk + 5)
        self.close_win.grid(column=5, row=kk + 5)


class main_menu:
    def __init__(self, file_data):
        self.Win_main_menu = Tk()
        self.file = file_data
        self.menu = menu_bar(self.Win_main_menu, self.file, main_menu)
        self.bill_menu_button = Button(self.Win_main_menu, text="BILL MENU", font=("ARIAL", 20, "bold"),
                                       command=self.go_to_bil_menu)
        self.bill_menu_button.grid(column=0, row=0)
        self.Win_main_menu.mainloop()

    def go_to_bil_menu(self):
        self.Win_main_menu.destroy()
        bill_menu(self.file)


class menu_bar:
    def __init__(self, Win, file, curent_window):
        self.file_data = file
        self.load_file = load_workbook(self.file_data)
        self.sheets = self.load_file.sheetnames
        self.curentwidow = curent_window
        print("hello")
        self.Win = Win
        self.menubar = Menu(self.Win)
        self.Win.config(menu=self.menubar)
        self.menubar.add_command(label='\u2302', command=self.go_to_main_menu)
        self.menubar.add_command(label='🔃', command=self.refresh)
        self.open_file = Menu(self.menubar, tearoff=0)
        self.menubar.add_cascade(label="open", menu=self.open_file)
        self.open_file.add_command(label=f"open:{self.sheets[0], self.sheets[1]} data base",
                                   command=lambda: startfile(self.file_data))
        self.open_file.add_separator()

    def go_to_main_menu(self):
        self.Win.destroy()
        main_menu(self.file_data)

    def refresh(self):
        self.Win.destroy()
        # self.Win.menu.destroy()
        self.curentwidow(self.file_data)


file = r"DATA.xlsx"
main_menu(file)
# "🏠"
