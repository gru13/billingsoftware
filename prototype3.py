from tkinter import *
from openpyxl import load_workbook
from datetime import datetime

file = "SAREE DATABASE.xlsx"


# Accessing data from excel speared sheet by using openpyxl.load_workbook
#
#
def main():
    file_ = "SAREE DATABASE.xlsx"
    global total_price
    total_price = 0
    workbook = load_workbook(file_)
    total_database = workbook['total data in out']
    billing_database = workbook['billing contain']
    saree_num_list_all = [str(a.value) for a in total_database["A"] if a.value is not None]
    # print(saree_num_list_all)
    saree_num_list_unsold = [str(a) for a in range(1, len(saree_num_list_all) + 1) if
                             total_database["G" + str(a + 1)].value is None]
    global saree_bought_by_customer
    saree_bought_by_customer = []
    # print(saree_num_list_unsold)
    #
    bill_no = [a.value for a in billing_database["A"] if a.value != None]
    bill_no = bill_no[1:]
    print(bill_no)
    if len(bill_no) == 0:
        bill_no = 1
    else:
        bill_no = [int(a) for a in bill_no]
        bill_no = max(bill_no) + 1

    def nxt():
        Win.destroy()
        main()

    # below p is just a parameter with no use but it should be given
    def saree_entry(p):
        global saree_bought_by_customer
        sno = str(len(saree_bought_by_customer) + 1)
        saree_no = customer_saree_no.get()
        if saree_no in saree_num_list_unsold:
            customer_saree_no.delete(0, END)
            design = total_database[
                "F" + str(saree_num_list_all.index(saree_no) + 1)].value  # getting design from database
            colour = total_database[
                "E" + str(saree_num_list_all.index(saree_no) + 1)].value  # getting colour from database
            if colour is None: colour = "unmentioned"
            saree_num_list_unsold.remove(saree_no)
            saree_bought_by_customer += [[sno, saree_no, design, colour]]  # structure of list
            print([sno, saree_no, design, colour])
            Label(Win, text=sno, width=10).grid(column=0, row=4 + int(sno))
            Label(Win, text=saree_no, width=10).grid(column=1, row=4 + int(sno))
            Label(Win, text=design, width=10).grid(column=2, row=4 + int(sno))
            Label(Win, text=colour, width=10).grid(column=3, row=4 + int(sno))
            globals()["saree" + sno] = Entry(Win, width=10)
            globals()["saree" + sno].grid(column=4, row=4 + int(sno))

    # noinspection PyGlobalUndefined
    def bill():
        global saree_bought_by_customer
        global total_price
        for price in saree_bought_by_customer:
            sno = price[0]
            price.append(globals()["saree" + sno].get())
            total_price += int(globals()["saree" + sno].get())
        print(saree_bought_by_customer)
        # we know that (here 'abcd..' are the column of sheet total_database)
        # A:"SAREE NUMBER"	B:"PURCHASE FROM"	C:"PURCHASE DATE"	D:"PURCHASE PRICE"	E:"COLOUR" F:"DESIGN" G:"SOLD PRICE"
        # H:"DATE OF SALE"	I:"CUSTOMER NAME"	J:"CUSTOMER NUM"	K:"ADDRESS"  L:"BILL NUMBER" N:"SNO IN BILL"
        for lst in saree_bought_by_customer:
            # now to append the data to total_database in workbook
            total_database["G" + str(saree_num_list_all.index(lst[1]) + 1)] = int(lst[-1])  # updating sold price
            total_database["H" + str(saree_num_list_all.index(lst[1]) + 1)] = datetime.now()  # DATE
            total_database["I" + str(saree_num_list_all.index(lst[1]) + 1)] = customer_name.get()  # CUS NAME
            total_database["J" + str(saree_num_list_all.index(lst[1]) + 1)] = int(customer_number.get())  # PHONE NUMBER
            total_database["K" + str(saree_num_list_all.index(lst[1]) + 1)] = customer_address.get()  # ADDRESS
            total_database["N" + str(saree_num_list_all.index(lst[1]) + 1)] = int(lst[0])  # SNO IN BIL
            total_database["L" + str(saree_num_list_all.index(lst[1]) + 1)] = bill_no
            # now to append the data to billing_database in workbook
            # A:"BILL NO" B:"SAREE NO" C:"DATE OF SALES" D:"CUSTOMER NAME" E:"DESIGN" F:"COLOUR" G:"SOLD PRICE"
            row_ = str(len(billing_database["A"]) + 1)
            billing_database["A" + row_] = bill_no
            billing_database["B" + row_] = lst[1]
            billing_database["C" + row_] = datetime.now()
            billing_database["D" + row_] = customer_number.get()
            billing_database["E" + row_] = lst[2]
            billing_database["F" + row_] = lst[3]
            billing_database["G" + row_] = int(lst[-1])
        workbook.save(file_)
        saree_bought_by_customer = []
        Label(Win, text="total=").grid(column=3, row=1000)
        Label(Win, text=total_price).grid(column=4, row=1000)
        Button(Win, text="next bill", command=nxt).grid(column=5, row=1000)

    # Create a gui window
    Win = Tk()
    Win.title("BILLING")
    # Win.iconbitmap("") for icon image adding
    #
    # Creating  all label widgets in window
    Label(Win, text="ABIRAMI SILKS", font=("ARIAL", 45)).grid(column=0, row=0, columnspan=10)
    Label(Win, text="Name").grid(column=0, row=1)
    Label(Win, text="Number").grid(column=2, row=1)
    Label(Win, text="Address").grid(column=0, row=2)
    Label(Win, text="Saree no").grid(column=0, row=3)
    #
    # Creating  all Entry widgets in window
    customer_name = Entry(Win, relief="sunken", justify=CENTER)
    customer_name.grid(column=1, row=1)
    customer_number = Entry(Win, relief="sunken", justify=CENTER)
    customer_number.grid(column=3, row=1)
    customer_address = Entry(Win, relief="sunken", justify=CENTER, width=45)
    customer_address.grid(column=1, row=2, columnspan=4)
    customer_saree_no = Entry(Win, relief="sunken", justify=CENTER)
    customer_saree_no.grid(column=1, row=3)
    customer_saree_no.bind("<Return>", saree_entry)
    Label(Win, text="sno", width=10).grid(row=4, column=0)
    Label(Win, text="saree no", width=10).grid(row=4, column=1)
    Label(Win, text="design", width=10).grid(row=4, column=2)
    Label(Win, text="colour", width=10).grid(row=4, column=3)
    Label(Win, text="price", width=10).grid(row=4, column=4)
    bill_button = Button(Win, text="bill", width=10, command=bill)
    bill_button.grid(column=5, row=1000)
    #
    Win.mainloop()


def logon_(P):
    if login.get() == "srb":
        main()
        root.destroy()


root = Tk()
Label(root, text="enter password to enter").grid(column=0, row=0)
login = Entry(root, )
login.bind("<Return>", logon_)
login.grid(column=1, row=0)
root.mainloop()
# pip install pyinstaller
# pyinstaller --onefile -w 'WORKING_PROTYPE_3_TEST.py'
# pyinstaller --onefile -w "working_protype_2_test_model.py"
